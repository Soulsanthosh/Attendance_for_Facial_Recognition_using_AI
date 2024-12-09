import face_recognition
import cv2
import numpy as np
import csv
import os
import glob
from datetime import datetime
import xlsxwriter as xl
import mysql.connector
import openpyxl as xl
from openpyxl import Workbook
import random as ra



video_capture=cv2.VideoCapture(0)


Nithin_image=face_recognition.load_image_file('512220104023.jpeg')
Nithin_encoding=face_recognition.face_encodings (Nithin_image)[0]

Santhosh_image=face_recognition.load_image_file('512220104028.jpg')
Santhosh_encoding=face_recognition.face_encodings (Santhosh_image)[0]

Kalaimamani_image=face_recognition.load_image_file('512220104013.jpeg')
Kalaimamani_encoding=face_recognition.face_encodings (Kalaimamani_image)[0]

Lokesh_image=face_recognition.load_image_file('512220104017.jpg')
Lokesh_encoding=face_recognition.face_encodings (Lokesh_image)[0]

known_face_encding=[Nithin_encoding,Santhosh_encoding,Kalaimamani_encoding,Lokesh_encoding]
known_face_names=["Nithin_J","Santhosh_S","Kalaimamani_S",'Lokesh_K']

pre_names={"Nithin_J":512220104023,"Santhosh_S":512220104028,"Kalaimamani_S":512220104013,'Lokesh_K':512220104017}
reg=[512220104023,512220104028,512220104013,512220104017]
mobile_no={512220104023:7305425420,512220104028:1234567891,512220104013:1234567890,512220104017:1234567890}
std_class={512220104023:'4',512220104028:'3',512220104013: '2',512220104017:'1'}
Dept={512220104023:'CSE',512220104028:'CSE',512220104013: 'IT',512220104017:'IT'}
student=known_face_names.copy()
face_location=[]
face_enconding=[]
face_names=[]
s=True
empty=set()
now= datetime.now()
current_date=now.strftime("%Y-%m-%d")

#f=open(current_date+'.csv','w+',newline='')
#lnwriter=csv.writer(f)
n=0
b=["n"]
absent=[]

while True:
    __,frame=video_capture.read()
    small_frame= cv2.resize(frame,(0,0),fx=0.25,fy=0.25)
    rgb_small_frame= small_frame[:,:,::-1]
    if True:
        face_locations=face_recognition.face_locations(rgb_small_frame)
        face_encondings=face_recognition.face_encodings(rgb_small_frame,face_locations)
        face_names=[]
        for face_encoding in face_encondings:
            matchs= face_recognition.compare_faces(known_face_encding,face_encoding)
            name=''
            face_distance=face_recognition.face_distance(known_face_encding,face_encoding)
            best_match_index=np.argmin(face_distance)
            if matchs[best_match_index]:
                name=known_face_names[best_match_index]
            face_names.append(name)

            if name in known_face_names:
                # print(name)
                empty.add(name)

    cv2.imshow("attendance",(frame))
    if cv2.waitKey(1) & 0xFF==ord('q'):   
        break
video_capture.release()
cv2.destroyAllWindows()    
n=n+1
# print(empty)
w=Workbook()
sheet1=w.active
sheet_name=['Absent_students']
worksheet=w.create_sheet(title='Absent_students')
worksheet.cell(row=2,column=2).value='Reg.No'
worksheet.cell(row=2,column=3).value='Name'
worksheet.cell(row=2,column=4).value='Year'
worksheet.cell(row=2,column=5).value='Department'
worksheet.cell(row=2,column=6).value='Mobile.No'

p=[pre_names[na] for na in empty]
    
s={i for i in (Dept.values())}
# print(s)
for i in range(1,len(std_class)+1):
    worksheet=w.create_sheet(title=(str(i)))
    sheet_name.append((str(i)))
    worksheet.cell(row=2,column=2).value='Reg.No'
    worksheet.cell(row=2,column=3).value='Name'
    worksheet.cell(row=2,column=4).value='Year'
    worksheet.cell(row=2,column=5).value='Department'
    worksheet.cell(row=2,column=6).value='Mobile.No'
r=3
ss=3
for name in known_face_names:
    # for i in known_face_names:
        n=pre_names[name]
        b=std_class[n] 
        if name in empty:
            worksheet=w[b]
            worksheet.cell(row=ss,column=2).value=n
            worksheet.cell(row=ss,column=3).value=name
            worksheet.cell(row=ss,column=4).value=std_class[n]
            worksheet.cell(row=ss,column=5).value=Dept[n]
            worksheet.cell(row=ss,column=6).value=mobile_no[n]
            ss+=1
            if len(empty)>ss:
                ss=3
        else:
            worksheet=w['Absent_students']
            worksheet.cell(row=r,column=2).value=n
            worksheet.cell(row=r,column=3).value=name
            worksheet.cell(row=r,column=4).value=std_class[n]
            worksheet.cell(row=r,column=5).value=Dept[n]
            worksheet.cell(row=r,column=6).value=mobile_no[n]
            r+=1
    # r+=1

# sheet = w['sheet 1']
print('Completed...')
w.save(f'{current_date}.xlsx')